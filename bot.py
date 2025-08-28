
import os
import html
import logging
import re
from datetime import datetime
from datetime import timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

DATA_DIR = Path('data')
DATA_DIR.mkdir(exist_ok=True)

load_dotenv()
BOT_TOKEN = os.getenv('BOT_TOKEN') or ''
TEST_MODE = (os.getenv('TEST_MODE') or '').strip() in {'1', 'true', 'True', 'yes', 'on'}

# Determine local timezone (fallback to system local if available)
try:
    LOCAL_TZ = datetime.now().astimezone().tzinfo
except Exception:
    LOCAL_TZ = None

# Basic logging to observe incoming updates/attachments in runtime
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s %(levelname)s %(name)s: %(message)s'
# )
# logger = logging.getLogger(__name__)

ASSESSMENT_SHEET = 'Оценивание'
ASSIGNMENTS_SHEET = 'Задания'
INFO_SHEET = 'Инфо'

# Fallback English sheet names if the user follows the screenshots exactly
SHEET_ALIASES = {
    ASSESSMENT_SHEET: {ASSESSMENT_SHEET, 'Assessment', 'Оценка', 'Оценивание', 'Оценки'},
    ASSIGNMENTS_SHEET: {ASSIGNMENTS_SHEET, 'Assignments', 'Задания', 'Дедлайны'},
    INFO_SHEET: {INFO_SHEET, 'Info', 'Информация'},
}

CourseWeights = Dict[str, float]
Assignment = Tuple[str, datetime, str]


def find_sheet_name(wb, canonical: str) -> Optional[str]:
    aliases = SHEET_ALIASES[canonical]
    for name in wb.sheetnames:
        if name in aliases:
            return name
    return None


def parse_weights(xlsx_path: Path) -> CourseWeights:
    wb = load_workbook(xlsx_path, data_only=True)
    name = find_sheet_name(wb, ASSESSMENT_SHEET)
    if not name:
        raise ValueError('Sheet with course weights not found')
    ws = wb[name]
    weights: CourseWeights = {}
    # Expect two columns: label and weight (like in screenshot)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or (row[0] is None and row[1] is None):
            continue
        label = str(row[0]).strip() if row[0] is not None else ''
        val = row[1]
        if label and isinstance(val, (int, float)):
            weights[label] = float(val)
    return weights


def parse_assignments(xlsx_path: Path) -> List[Assignment]:
    wb = load_workbook(xlsx_path, data_only=True)
    name = find_sheet_name(wb, ASSIGNMENTS_SHEET)
    if not name:
        raise ValueError('Sheet with assignments not found')
    ws = wb[name]
    results: List[Assignment] = []
    # Expect columns: Title, Due date, Link
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or (row[0] is None and row[1] is None and row[2] is None):
            continue
        title = (str(row[0]).strip()) if row[0] is not None else ''
        due_raw = row[1]
        link = (str(row[2]).strip()) if row[2] is not None else ''
        if not title:
            continue
        # Convert date: accept datetime or dd.mm.yyyy string
        if isinstance(due_raw, datetime):
            due = due_raw
        elif isinstance(due_raw, (int, float)):
            # Excel serial date, attempt conversion
            try:
                from openpyxl.utils.datetime import from_excel
                due = from_excel(due_raw)
            except Exception:
                continue
        elif isinstance(due_raw, str):
            m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", due_raw.strip())
            if not m:
                continue
            day, month, year = map(int, m.groups())
            due = datetime(year, month, day)
        else:
            continue
        results.append((title, due, link))
    # sort by date
    results.sort(key=lambda x: x[1])
    return results


def format_formula(weights: CourseWeights) -> str:
    if not weights:
        return 'Формула пока не задана. Загрузите Excel с листом "Оценивание".'
    parts = [f"{name}×{w:g}" for name, w in weights.items()]
    return 'Итог = ' + ' + '.join(parts)


def format_nearest(assignments: List[Assignment], limit: int = 5) -> str:
    if not assignments:
        return 'Нет данных о дедлайнах. Загрузите Excel с листом "Задания".'
    now = datetime.now()
    horizon = now + timedelta(days=14)
    window = [a for a in assignments if now <= a[1] <= horizon]
    lines = []
    for title, due, link in window[:limit]:
        date_str = due.strftime('%d.%m.%Y')
        safe_title = html.escape(title)
        if link:
            safe_link = html.escape(link)
            lines.append(f"• <a href=\"{safe_link}\">{safe_title}</a> — {date_str}")
        else:
            lines.append(f"• {safe_title} — {date_str}")
    if not lines:
        return 'В ближайшие 2 недели дедлайнов нет.'
    return "\n".join(lines)


def get_chat_file(chat_id: int) -> Path:
    return DATA_DIR / f"{chat_id}.xlsx"


def parse_info(xlsx_path: Path) -> List[Tuple[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    name = find_sheet_name(wb, INFO_SHEET)
    if not name:
        raise ValueError('Лист с информацией ("Инфо") не найден')
    ws = wb[name]
    rows: List[Tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or (row[0] is None and row[1] is None):
            continue
        key = str(row[0]).strip() if row[0] is not None else ''
        val = str(row[1]).strip() if row[1] is not None else ''
        if key:
            rows.append((key, val))
    return rows


def format_info(items: List[Tuple[str, str]]) -> str:
    if not items:
        return 'Нет данных на листе "Инфо".'
    username_keys = {'Преподаватель', 'Ассистент', 'Канал'}
    lines: List[str] = ['Привет! Ниже приведены ссылки на основные ресурсы курса 👇']
    for key, val in items:
        safe_key = html.escape(key)
        if key in username_keys:
            uname = re.sub(r"^@+", '', val).strip()
            if uname:
                display = '@' + uname
            else:
                display = ''
            lines.append(f"• <b>{safe_key}</b>: {html.escape(display)}")
        else:
            link = val.strip()
            # Basic normalization: add scheme if missing for tg client to open
            if link and not re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", link):
                link = 'https://' + link
            safe_link = html.escape(link)
            lines.append(f"• <a href=\"{safe_link}\">{safe_key}</a>")
    return "\n\n".join(lines)


async def daily_reminder_callback(context: ContextTypes.DEFAULT_TYPE) -> None:
    data = context.job.data or {}
    chat_id: int = data.get('chat_id')
    xlsx_path: str = data.get('xlsx_path')
    if not chat_id or not xlsx_path:
        return
    try:
        assignments = parse_assignments(Path(xlsx_path))
    except Exception:
        return

    now = datetime.now(tz=LOCAL_TZ) if LOCAL_TZ else datetime.now()
    today = now.date()
    week_target = today + timedelta(days=7)
    day_target = today + timedelta(days=1)

    week_items: List[Assignment] = []
    day_items: List[Assignment] = []
    for title, due, link in assignments:
        due_date = (due.astimezone(LOCAL_TZ).date() if (LOCAL_TZ and due.tzinfo) else due.date())
        if due_date == week_target:
            week_items.append((title, due, link))
        if due_date == day_target:
            day_items.append((title, due, link))

    def build_message(label: str, items: List[Assignment]) -> Optional[str]:
        if not items:
            return None
        lines: List[str] = [f"🔔 Напоминание <b>{label}</b> до дедлайна:"]
        for title, due, link in items:
            date_str = due.strftime('%d.%m.%Y')
            safe_title = html.escape(title)
            if link:
                safe_link = html.escape(link)
                lines.append(f"• <a href=\"{safe_link}\">{safe_title}</a> — {date_str}")
            else:
                lines.append(f"• {safe_title} — {date_str}")
        return "\n".join(lines)

    week_text = build_message('за неделю', week_items)
    day_text = build_message('за день', day_items)

    if week_text:
        await context.bot.send_message(chat_id=chat_id, text=week_text, parse_mode=ParseMode.HTML, disable_web_page_preview=True)
    if day_text:
        await context.bot.send_message(chat_id=chat_id, text=day_text, parse_mode=ParseMode.HTML, disable_web_page_preview=True)


def schedule_chat_reminders(chat_id: int, xlsx_path: Path, context: ContextTypes.DEFAULT_TYPE) -> None:
    jq = context.job_queue
    if jq is None:
        return
    # Remove previous jobs for this chat
    jq.scheduler.remove_all_jobs()

    if TEST_MODE:
        # Run every 15 seconds for quick verification
        jq.run_repeating(
            daily_reminder_callback,
            interval=15,
            first=0,
            data={'chat_id': chat_id, 'xlsx_path': str(xlsx_path)},
            name=f"daily:{chat_id}"
        )
        return

    # Schedule daily at 10:00 local time
    tzinfo = LOCAL_TZ if LOCAL_TZ else ZoneInfo('UTC')
    run_time = datetime.now(tzinfo).replace(hour=10, minute=0, second=0, microsecond=0).timetz()
    jq.run_daily(
        daily_reminder_callback,
        time=run_time,
        data={'chat_id': chat_id, 'xlsx_path': str(xlsx_path)},
        name=f"daily:{chat_id}"
    )


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    xlsx_path = get_chat_file(chat_id)
    if not xlsx_path.exists():
        await update.message.reply_text(
            'Сначала загрузите Excel с листами "Оценивание" и "Задания". '
            'Отправьте файл прямо сюда.'
        )
        return
    try:
        weights = parse_weights(xlsx_path)
        assignments = parse_assignments(xlsx_path)
    except Exception as exc:
        await update.message.reply_text(
            'Не удалось прочитать файл. Загрузите корректный Excel. '
            f'Ошибка: {exc}'
        )
        return

    formula = format_formula(weights)
    deadlines = format_nearest(assignments)

    text = (
        '<b>Формула оценки</b>\n'
        f'{formula}\n\n'
        '<b>Ближайшие дедлайны</b>\n'
        f'{deadlines}'
    )
    await update.message.reply_html(text)


async def info_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    xlsx_path = get_chat_file(chat_id)
    if not xlsx_path.exists():
        await update.message.reply_text(
            'Сначала загрузите Excel с листом "Инфо". Отправьте файл прямо сюда.'
        )
        return
    try:
        items = parse_info(xlsx_path)
    except Exception as exc:
        await update.message.reply_text(
            'Не удалось прочитать лист "Инфо". Загрузите корректный Excel. '
            f'Ошибка: {exc}'
        )
        return
    text = format_info(items)
    await update.message.reply_html(text)

async def update_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    
    # Check if this is the authorized user
    if chat_id != 669636800:
        await update.message.reply_text(
            'У вас нет прав для использования этой команды. '
            'Доступ ограничен для администратора бота.'
        )
        return
    
    # Set a flag in context to indicate this is an update request
    context.user_data['expecting_update'] = True
    
    await update.message.reply_text(
        'Отправьте новый Excel-файл для обновления информации. '
        'Файл должен содержать листы "Оценивание", "Задания" и "Инфо".'
    )

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    is_update_request = context.user_data.get('expecting_update', False)

    message = update.message
    if not message:
        return

    # Prefer explicit document field
    doc = message.document
    # Fallback: sometimes attachments come via effective_attachment list
    if not doc and message.effective_attachment:
        attachments = message.effective_attachment
        if isinstance(attachments, list):
            for att in attachments:
                file_name_attr = getattr(att, 'file_name', None)
                if file_name_attr and str(file_name_attr).lower().endswith(('.xlsx', '.xlsm')):
                    doc = att
                    break

    # logger.info("Incoming attachment: %s", getattr(doc, 'file_name', None))
    if not doc:
        # Not an Excel document; ignore silently to avoid noise
        return
    file_name = (doc.file_name or '').lower()
    if not (file_name.endswith('.xlsx') or file_name.endswith('.xlsm')):
        await update.message.reply_text('Пожалуйста, отправьте Excel-файл (.xlsx или .xlsm).')
        return

    if is_update_request:
        await update.message.reply_text('Обновляю информацию из нового файла...')
    else:
        await update.message.reply_text('Получаю файл...')
    
    new_path = get_chat_file(chat_id)
    tg_file = await doc.get_file()
    await tg_file.download_to_drive(str(new_path))

    # Validate content
    try:
        _ = parse_weights(new_path)
        _ = parse_assignments(new_path)
    except Exception as exc:
        await update.message.reply_text(
            'Файл сохранён, но не удалось его прочитать полностью. '
            f'Проверьте структуру листов. Ошибка: {exc}'
        )
        return

    # Clear the update flag
    context.user_data.pop('expecting_update', None)

    if is_update_request:
        await update.message.reply_text('Информация успешно обновлена! Используйте /help для сводки.')
    else:
        await update.message.reply_text('Файл сохранён для этого чата. Используйте /help для сводки.')

    # Schedule aggregated daily reminders for this chat
    try:
        schedule_chat_reminders(chat_id, new_path, context)
        if TEST_MODE:
            await update.message.reply_text('Тестовые напоминания активированы (каждые 15 секунд).')
        else:
            await update.message.reply_text('Ежедневные напоминания запланированы на 10:00.')
    except Exception as exc:
        await update.message.reply_text(f'Не удалось запланировать напоминания: {exc}')


def main() -> None:
    token = BOT_TOKEN
    if not token:
        raise SystemExit('Please set BOT_TOKEN in .env')
    # Define startup hook to pre-schedule reminders for existing chats
    async def startup(application: Application) -> None:
        for path in DATA_DIR.glob('*.xlsx'):
            try:
                chat_id = int(path.stem)
            except ValueError:
                continue
            schedule_chat_reminders(chat_id, path, application)

    app = Application.builder().token(token).post_init(startup).build()

    app.add_handler(CommandHandler('help', help_cmd))
    app.add_handler(CommandHandler('info', info_cmd))
    app.add_handler(CommandHandler('update', update_cmd))
    # Handle standard documents and any kind of attachments (covers some clients)
    app.add_handler(MessageHandler(filters.Document.ALL | filters.ATTACHMENT, handle_document))

    app.run_polling()
    print('Bot is running')


if __name__ == '__main__':
    main()
